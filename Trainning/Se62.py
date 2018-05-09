# -*- coding: UTF-8 -*-
import time
import os
import xlrd
from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
from xlutils.copy import copy
from openpyxl import load_workbook
from openpyxl import Workbook

def r_excel(name,r,c):
    wb=load_workbook(name)
    sheet = wb.active
    return sheet.cell(row=r, column=c).value

def w_excel(name,r,c,v):
    wb=load_workbook(name)
    sheet = wb.active
    sheet.cell(row=r, column=c).value=v
    wb.save(name)

iedriver ="C:\Program Files (x86)\Internet Explorer\IEDriverServer.exe"
os.environ["webdriver.ie.driver"] = iedriver

if __name__=="__main__":
    for a in range(2, 7):  # 2-7
        driver = webdriver.Ie(iedriver)
        driver.get("http://www.baidu.com")
        print type(r_excel("case.xlsx", a, 3))
        # var = str(r_excel("case.xlsx", a, 3))
        if a == 6:
            driver.find_element_by_id("kw").send_keys(str(r_excel("case.xlsx", a, 3)))
        else:
            driver.find_element_by_id("kw").send_keys(r_excel("case.xlsx", a, 3))
        driver.find_element_by_id("su").click()
        time.sleep(2)
        print driver.title
        w_excel("case.xlsx", a, 5, driver.title)
        driver.quit()

#---------------------------------------------------------------------------------------
    # driver.find_element_by_class_name("s_ipt").send_keys("Selenium")
    #driver.find_element_by_name("wd").send_keys("Selenium")
    #driver.find_element_by_link_text("About  Baidu").click()
    #driver.find_element_by_partial_link_text("About").click()
    #driver.find_element_by_tag_name("body").send_keys(Keys.F11)
    # driver.find_element_by_css_selector("#kw").send_keys("Selenium")
    #driver.find_element_by_xpath("//*[@id='kw']").send_keys("Selenium")
    #driver.implicitly_wait(3)
    #WebDriverWait(driver, 10).until(lambda x: x.find_element_by_id("kw").send_keys("Selenium"))
    #driver.find_element_by_link_text("登录").click()
    #driver.find_element_by_name("tj_login").click()
    # element0 = driver.find_elements_by_name("tj_login")
    # for ele0 in element0:
    #     if ele0.is_displayed():
    #         ele0.click()
    # driver.find_element_by_id("TANGRAM__PSP_10__footerULoginBtn").click()
    #
    # element1 = driver.find_element_by_class_name("tang-content")
    # element11 = element1.find_element_by_id("TANGRAM__PSP_10__userName")
    # element11.clear()
    # element11.send_keys("aaa")
    #
    # element2 = element1.find_element_by_id("TANGRAM__PSP_10__password")
    # element2.clear()
    # element2.send_keys("bbb")
    #####################################################################################################
    # driver.get("http://www.qq.com") # 访问腾讯首页
    #
    # loginLayout = driver.find_element_by_id("loginGrayLayout")  # 根据id定位“登录”按钮
    # loginLayout.click()  # 点击“登录”按钮，会弹出一个IFrame
    #
    # driver.switch_to_frame("login_frame")  # 将焦点移动到IFrame上
    #
    # loginByAccount = driver.find_element_by_id("switcher_plogin")  # 根据id定位“用户名密码登录”按钮
    # loginByAccount.click()  # 点一下这个按钮

    #driver.quit()  # 退出浏览器
    # driver.find_element_by_name("wd").send_keys("51testing")
    #
    # driver.find_element_by_id("su").submit()
    # time.sleep(2)
    # print driver.title
    # ver("123",driver.title)
    #driver.quit()